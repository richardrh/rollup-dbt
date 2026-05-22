from __future__ import annotations

import builtins
from collections.abc import Iterator
from pathlib import Path

import polars as pl
from openpyxl import Workbook

from rollup import cli


def _write_minimal_verisk_workbook(path: Path, *, aal: float = 12.5) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PML by LOB"
    for column, value in enumerate(
        ["Analysis", "ExposureAttribute", "CatalogTypeCode", "aal_0", "aep_100"],
        start=1,
    ):
        worksheet.cell(row=7, column=column, value=value)
    for column, value in enumerate(
        ["TEST_PERIL", "TEST_LOB", "STC", aal, 1000.0],
        start=1,
    ):
        worksheet.cell(row=8, column=column, value=value)
    workbook.save(path)


def _mock_input(monkeypatch, responses: list[str]) -> None:
    response_iter: Iterator[str] = iter(responses)
    monkeypatch.setattr(builtins, "input", lambda prompt="": next(response_iter))


def test_generate_ep_summaries_interactive_selects_file_and_overwrites_output(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    data_root = tmp_path / "data"
    workbook_path = data_root / "ep_summaries" / "verisk" / "selected.xlsx"
    output_path = data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv"
    _write_minimal_verisk_workbook(workbook_path)
    output_path.write_text("old contents\n", encoding="utf-8")
    _mock_input(monkeypatch, ["1", "1", "y"])

    exit_code = cli.main(["--data-root", str(data_root), "generate-ep-summaries"])

    assert exit_code == 0
    output = pl.read_csv(output_path)
    assert output.columns == [
        "vendor",
        "analysis_id",
        "modelled_lob",
        "modelled_peril",
        "ep_type",
        "return_period",
        "loss",
    ]
    assert output.to_dicts() == [
        {
            "vendor": "verisk",
            "analysis_id": "TEST_PERIL",
            "modelled_lob": "TEST_LOB",
            "modelled_peril": "TEST_PERIL",
            "ep_type": "AAL",
            "return_period": 0,
            "loss": 12.5,
        },
        {
            "vendor": "verisk",
            "analysis_id": "TEST_PERIL",
            "modelled_lob": "TEST_LOB",
            "modelled_peril": "TEST_PERIL",
            "ep_type": "AEP",
            "return_period": 100,
            "loss": 1000.0,
        },
    ]
    captured = capsys.readouterr().out
    assert "Select EP summary vendor:" in captured
    assert "Select source XLSX workbook:" in captured
    assert "EP summary written to" in captured


def test_generate_ep_summaries_non_interactive_accepts_xlsx_filename(
    tmp_path: Path,
    monkeypatch,
) -> None:
    data_root = tmp_path / "data"
    workbook_path = data_root / "ep_summaries" / "verisk" / "selected.xlsx"
    output_path = data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv"
    _write_minimal_verisk_workbook(workbook_path, aal=25.0)
    monkeypatch.setattr(
        builtins,
        "input",
        lambda prompt="": (_ for _ in ()).throw(AssertionError("input should not be called")),
    )

    exit_code = cli.main(
        [
            "--data-root",
            str(data_root),
            "generate-ep-summaries",
            "--vendor",
            "verisk",
            "--xlsx",
            "selected.xlsx",
            "--yes",
        ]
    )

    assert exit_code == 0
    assert pl.read_csv(output_path).item(0, "loss") == 25.0


def test_generate_ep_summaries_explicit_args_without_yes_skip_prompt_for_new_output(
    tmp_path: Path,
    monkeypatch,
) -> None:
    data_root = tmp_path / "data"
    workbook_path = data_root / "ep_summaries" / "verisk" / "selected.xlsx"
    output_path = data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv"
    _write_minimal_verisk_workbook(workbook_path, aal=50.0)
    monkeypatch.setattr(
        builtins,
        "input",
        lambda prompt="": (_ for _ in ()).throw(AssertionError("input should not be called")),
    )

    exit_code = cli.main(
        [
            "--data-root",
            str(data_root),
            "generate-ep-summaries",
            "--vendor",
            "verisk",
            "--xlsx",
            "selected.xlsx",
        ]
    )

    assert exit_code == 0
    assert pl.read_csv(output_path).item(0, "loss") == 50.0


def test_generate_ep_summaries_returns_nonzero_when_input_ends(
    tmp_path: Path,
    monkeypatch,
    capsys,
) -> None:
    data_root = tmp_path / "data"
    workbook_path = data_root / "ep_summaries" / "verisk" / "selected.xlsx"
    _write_minimal_verisk_workbook(workbook_path)
    monkeypatch.setattr(
        builtins,
        "input",
        lambda prompt="": (_ for _ in ()).throw(EOFError),
    )

    exit_code = cli.main(["--data-root", str(data_root), "generate-ep-summaries"])

    assert exit_code == 1
    assert "Input ended before EP summary generation could continue." in capsys.readouterr().err


def test_generate_ep_summaries_confirmation_defaults_to_no_without_overwrite(
    tmp_path: Path,
    monkeypatch,
) -> None:
    data_root = tmp_path / "data"
    workbook_path = data_root / "ep_summaries" / "verisk" / "selected.xlsx"
    output_path = data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv"
    _write_minimal_verisk_workbook(workbook_path)
    output_path.write_text("old contents\n", encoding="utf-8")
    _mock_input(monkeypatch, ["1", "1", ""])

    exit_code = cli.main(["--data-root", str(data_root), "generate-ep-summaries"])

    assert exit_code == 0
    assert output_path.read_text(encoding="utf-8") == "old contents\n"
