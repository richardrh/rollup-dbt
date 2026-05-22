from __future__ import annotations

from pathlib import Path

import polars as pl
from openpyxl import Workbook

from rollup.ep_summary_generator import (
    build_verisk_ep_summary,
    generate_vendor_ep_summary,
    scan_ep_summary_workbooks,
)


def _write_minimal_verisk_workbook(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PML by LOB"
    for column, value in enumerate(
        ["Analysis", "ExposureAttribute", "CatalogTypeCode", "aal_0", "aep_100"],
        start=1,
    ):
        worksheet.cell(row=7, column=column, value=value)
    for column, value in enumerate(
        ["TEST_PERIL", "TEST_LOB", "STC", 12.5, 1000.0],
        start=1,
    ):
        worksheet.cell(row=8, column=column, value=value)
    workbook.save(path)


def test_build_verisk_ep_summary_uses_explicit_workbook_path(tmp_path: Path) -> None:
    workbook_path = tmp_path / "custom source.xlsx"
    _write_minimal_verisk_workbook(workbook_path)

    frame = build_verisk_ep_summary(workbook_path)

    assert frame.to_dicts() == [
        {
            "vendor": "verisk",
            "analysis_id": "TEST_PERIL",
            "modelled_lob": "TEST_LOB",
            "modelled_peril": "TEST_PERIL",
            "ep_type": "AAL",
            "return_period": 0,
            "loss": 12.5,
        },
        {
            "vendor": "verisk",
            "analysis_id": "TEST_PERIL",
            "modelled_lob": "TEST_LOB",
            "modelled_peril": "TEST_PERIL",
            "ep_type": "AEP",
            "return_period": 100,
            "loss": 1000.0,
        },
    ]


def test_scan_workbooks_sorts_and_generate_vendor_writes_canonical_output(
    tmp_path: Path,
) -> None:
    data_root = tmp_path / "data"
    source_dir = data_root / "ep_summaries" / "verisk"
    first_workbook = source_dir / "b.xlsx"
    second_workbook = source_dir / "A.xlsx"
    _write_minimal_verisk_workbook(first_workbook)
    _write_minimal_verisk_workbook(second_workbook)

    assert scan_ep_summary_workbooks(data_root, "verisk") == [second_workbook, first_workbook]

    output_path = generate_vendor_ep_summary(data_root, "verisk", second_workbook)

    assert output_path == source_dir / "verisk_ep_summary.long.csv"
    output = pl.read_csv(output_path)
    assert output.select("vendor", "analysis_id", "loss").to_dicts() == [
        {"vendor": "verisk", "analysis_id": "TEST_PERIL", "loss": 12.5},
        {"vendor": "verisk", "analysis_id": "TEST_PERIL", "loss": 1000.0},
    ]
