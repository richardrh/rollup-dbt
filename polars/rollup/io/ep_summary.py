"""Convert wide EP-summary xlsx exports into canonical long-format CSVs.

RiskLink source: 'OEPAEP Curves' sheet. Title decoration in rows 1-6, header
at row 7, data from row 8 onward. Wide layout has one column per
(ep_type, return_period): AAL, OEP_2, OEP_5, ..., AEP_2, ...

Verisk source: 'PML by LOB' sheet. Header at row 7, data from row 8 onward.
Wide layout has aal_0.0, aep_<rp>.0, and oep_<rp>.0 columns.

Long format: one row per
``(vendor, analysis_id, modelled_lob, modelled_peril, ep_type, return_period, loss)``.
- ep_type in {'AAL', 'OEP', 'AEP'}
- return_period = 0 for AAL rows, otherwise the return-period integer (2, 5, 10, ...)
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import polars as pl

from rollup.config import VendorName
from rollup.schemas.columns import CanonicalEpSummaryCol as EP


_HEADER_ROW = 7   # 1-indexed default for GC exports with prefixed RP columns
_DATA_START_ROW = 8
_EP_SHEET = "OEPAEP Curves"
_VERISK_PML_SHEET = "PML by LOB"

# Strict — only OEP_<int> / AEP_<int> are accepted as RP columns.
# A header like 'OEP_DIFF' or 'AEP_TOTAL' is silently ignored, not parsed
# as a return period (which would crash int() and skip the whole file).
_RP_COLUMN = re.compile(r"^(?P<kind>OEP|AEP)_(?P<rp>\d+)$")
_VERISK_RP_COLUMN = re.compile(r"^(?P<kind>oep|aep)_(?P<rp>\d+(?:\.0+)?)$", re.IGNORECASE)


def _clean(v):
    """Return None for blank/empty/dash values; otherwise return v unchanged."""
    if v is None:
        return None
    if isinstance(v, str):
        stripped = v.strip()
        return None if stripped == "" or stripped == "-" else stripped
    return v


def _clean_header(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        stripped = v.strip()
        return stripped or None
    return str(v).strip()


def _find_header_row(rows: list[tuple], required: tuple[str, ...], path: Path, sheet: str) -> int:
    """Return the 0-indexed header row containing the required columns."""
    for idx, row in enumerate(rows[:20]):
        cleaned = {_clean_header(v) for v in row}
        if all(col in cleaned for col in required):
            return idx
    raise ValueError(f"{path}: missing required columns {required!r} in sheet {sheet!r}")


def read_risklink_ep_summary(path: Path) -> pl.DataFrame:
    """Read a RiskLink EP-summary xlsx into canonical EP-summary long format.

    Reads the 'OEPAEP Curves' sheet. Header row is row 7 (1-indexed); data
    begins at row 8. The source ``LOB`` is emitted as analyst-facing
    ``modelled_lob`` so dry-run can resolve it to ``lobs.csv``.

    Raises KeyError if the sheet is absent, ValueError if required columns are
    missing or there are fewer rows than expected.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_EP_SHEET]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    header_row = _find_header_row(rows, ("ID", "LOB", "RegionPeril", "AAL"), path, _EP_SHEET)
    header = list(rows[header_row])
    ep_type_header = list(rows[header_row - 1]) if header_row > 0 else []
    data = [list(r) for r in rows[header_row + 1:]]

    # Map column name -> index. Skip leading/trailing blank columns.
    col_idx: dict[str, int] = {
        cleaned: i
        for i, name in enumerate(header)
        if (cleaned := _clean_header(name)) is not None
    }

    for required in ("ID", "LOB", "RegionPeril", "AAL"):
        if required not in col_idx:
            raise ValueError(
                f"{path}: missing required column {required!r} in header row {header_row + 1}"
            )

    # Identify OEP_<int> / AEP_<int> columns; ignore non-numeric suffixes.
    rp_columns: list[tuple[int, int, str]] = []   # (col_idx, rp, ep_type)
    for idx, name in enumerate(header):
        c = _clean_header(name)
        if c is None:
            continue
        m = _RP_COLUMN.match(c)
        if m:
            rp_columns.append((idx, int(m["rp"]), m["kind"]))
            continue

        # Some RiskLink workbooks have one row of EP types (OEP/AEP) above a
        # header row where the RP columns are bare numbers: 2, 5, 10, ...
        if idx < len(ep_type_header):
            kind = str(ep_type_header[idx] or "").strip().upper()
            if kind in {"OEP", "AEP"}:
                try:
                    rp_columns.append((idx, int(float(c)), kind))
                except ValueError:
                    pass
    rp_columns.sort(key=lambda t: (t[2], t[1]))   # OEP first, then AEP, by RP

    out_rows: list[dict[str, int | float | str | None]] = []

    for row in data:
        # Skip entirely blank rows (trailing blank rows in the sheet).
        if all(
            c is None or (isinstance(c, str) and c.strip() == "") for c in row
        ):
            continue

        id_raw = _clean(row[col_idx["ID"]]) if col_idx["ID"] < len(row) else None
        if id_raw is None:
            continue
        try:
            id_int = int(id_raw)
        except (TypeError, ValueError):
            continue

        lob = _clean(row[col_idx["LOB"]]) if col_idx["LOB"] < len(row) else None
        region_peril = (
            _clean(row[col_idx["RegionPeril"]])
            if col_idx["RegionPeril"] < len(row)
            else None
        )

        # --- AAL row (rp = 0) ---
        aal_raw = _clean(row[col_idx["AAL"]]) if col_idx["AAL"] < len(row) else None
        if aal_raw is not None:
            try:
                out_rows.append(
                    {
                        EP.VENDOR:         VendorName.RISKLINK.value,
                        EP.ANALYSIS_ID:    str(id_int),
                        EP.MODELLED_LOB:   str(lob) if lob is not None else None,
                        EP.MODELLED_PERIL: str(region_peril) if region_peril is not None else None,
                        EP.EP_TYPE:        "AAL",
                        EP.RETURN_PERIOD:  0,
                        EP.LOSS:           float(aal_raw),
                    }
                )
            except (TypeError, ValueError):
                pass

        # --- OEP and AEP rows ---
        for idx, rp, kind in rp_columns:
            v = _clean(row[idx]) if idx < len(row) else None
            if v is None:
                continue
            try:
                out_rows.append(
                    {
                        EP.VENDOR:         VendorName.RISKLINK.value,
                        EP.ANALYSIS_ID:    str(id_int),
                        EP.MODELLED_LOB:   str(lob) if lob is not None else None,
                        EP.MODELLED_PERIL: str(region_peril) if region_peril is not None else None,
                        EP.EP_TYPE:        kind,
                        EP.RETURN_PERIOD:  rp,
                        EP.LOSS:           float(v),
                    }
                )
            except (TypeError, ValueError):
                continue

    return pl.DataFrame(
        out_rows,
        schema={
            EP.VENDOR:         pl.String,
            EP.ANALYSIS_ID:    pl.String,
            EP.MODELLED_LOB:   pl.String,
            EP.MODELLED_PERIL: pl.String,
            EP.EP_TYPE:        pl.String,
            EP.RETURN_PERIOD:  pl.Int64,
            EP.LOSS:           pl.Float64,
        },
    )


def read_verisk_ep_summary(path: Path) -> pl.DataFrame:
    """Read a Verisk EP-summary xlsx into canonical EP-summary long format.

    Reads the ``PML by LOB`` sheet. Header row is row 7 (1-indexed); data
    begins at row 8. The source layout is one row per
    ``(Analysis, ExposureAttribute, CatalogTypeCode)`` with wide columns such as
    ``aal_0.0``, ``aep_200.0`` and ``oep_200.0``. The output is one row per
    ``(analysis, modelled_lob, ep_type, return_period)``. Only STC catalogue
    rows are included so EP summaries align with Verisk YLT staging. Verisk
    workbooks do not carry numeric analysis IDs, so the raw ``Analysis`` label
    is emitted in both ``analysis_id`` and ``modelled_peril``; dry-run resolves
    analyst-selected numeric IDs through ``analyses.csv``.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[_VERISK_PML_SHEET]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    if len(rows) < _DATA_START_ROW:
        raise ValueError(
            f"{path}: '{_VERISK_PML_SHEET}' has fewer than {_DATA_START_ROW} rows"
        )

    header = list(rows[_HEADER_ROW - 1])
    data = [list(r) for r in rows[_DATA_START_ROW - 1:]]
    col_idx: dict[str, int] = {
        str(name).strip(): i
        for i, name in enumerate(header)
        if name is not None and str(name).strip()
    }

    for required in ("Analysis", "ExposureAttribute", "CatalogTypeCode", "aal_0.0"):
        if required not in col_idx:
            raise ValueError(
                f"{path}: missing required column {required!r} in header row {_HEADER_ROW}"
            )

    rp_columns: list[tuple[int, int, str]] = []
    for c in col_idx:
        m = _VERISK_RP_COLUMN.match(c)
        if m:
            rp_columns.append((col_idx[c], int(float(m["rp"])), m["kind"].upper()))
    rp_columns.sort(key=lambda t: (t[2], t[1]))

    out_rows: list[dict[str, int | float | str | None]] = []
    for row in data:
        if all(c is None or (isinstance(c, str) and c.strip() == "") for c in row):
            continue

        analysis = _clean(row[col_idx["Analysis"]]) if col_idx["Analysis"] < len(row) else None
        lob = (
            _clean(row[col_idx["ExposureAttribute"]])
            if col_idx["ExposureAttribute"] < len(row)
            else None
        )
        catalog_type = (
            _clean(row[col_idx["CatalogTypeCode"]])
            if col_idx["CatalogTypeCode"] < len(row)
            else None
        )
        if analysis is None or lob is None:
            continue
        if "STC" not in str(catalog_type or "").strip().upper():
            continue

        aal_raw = _clean(row[col_idx["aal_0.0"]]) if col_idx["aal_0.0"] < len(row) else None
        if aal_raw is not None:
            try:
                out_rows.append({
                    EP.VENDOR:         VendorName.VERISK.value,
                    EP.ANALYSIS_ID:    str(analysis),
                    EP.MODELLED_LOB:   str(lob),
                    EP.MODELLED_PERIL: str(analysis),
                    EP.EP_TYPE:        "AAL",
                    EP.RETURN_PERIOD:  0,
                    EP.LOSS:           float(aal_raw),
                })
            except (TypeError, ValueError):
                pass

        for idx, rp, kind in rp_columns:
            v = _clean(row[idx]) if idx < len(row) else None
            if v is None:
                continue
            try:
                out_rows.append({
                    EP.VENDOR:         VendorName.VERISK.value,
                    EP.ANALYSIS_ID:    str(analysis),
                    EP.MODELLED_LOB:   str(lob),
                    EP.MODELLED_PERIL: str(analysis),
                    EP.EP_TYPE:        kind,
                    EP.RETURN_PERIOD:  rp,
                    EP.LOSS:           float(v),
                })
            except (TypeError, ValueError):
                continue

    return pl.DataFrame(
        out_rows,
        schema={
            EP.VENDOR:         pl.String,
            EP.ANALYSIS_ID:    pl.String,
            EP.MODELLED_LOB:   pl.String,
            EP.MODELLED_PERIL: pl.String,
            EP.EP_TYPE:        pl.String,
            EP.RETURN_PERIOD:  pl.Int64,
            EP.LOSS:           pl.Float64,
        },
    )


def write_long_csv(df: pl.DataFrame, output: Path) -> None:
    """Write `df` as CSV to `output`, creating parent directories as needed."""
    output.parent.mkdir(parents=True, exist_ok=True)
    df.write_csv(output)


def convert_ep_summaries_to_csv(ep_dir: Path, vendor: VendorName) -> list[Path]:
    """For each xlsx under `ep_dir`, write a sibling `<stem>.long.csv`.

    Returns the list of CSV paths written. Skips xlsx files that don't have
    an 'OEPAEP Curves' sheet (e.g. portfolio-list workbooks) and any xlsx that
    raises a ValueError during parsing.
    """
    written: list[Path] = []
    if not ep_dir.exists():
        return written

    for xlsx in sorted(ep_dir.glob("*.xlsx")):
        try:
            if vendor == VendorName.RISKLINK:
                df = read_risklink_ep_summary(xlsx)
            else:
                df = read_verisk_ep_summary(xlsx)
        except (KeyError, ValueError):
            continue

        out = xlsx.with_name(xlsx.stem + ".long.csv")
        write_long_csv(df, out)
        written.append(out)

    return written
