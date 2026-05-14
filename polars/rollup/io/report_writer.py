"""CSV + xlsx writers for the end-of-run EP summary report.

The CSV is the source of truth: long-format, greppable, diffable in PRs. The
xlsx is a formatted view derived from the same DataFrame — one sheet per grain
(total / rollup_lob / peril), each sheet pivoted so variants are rows and
(ep_type, rp) columns. Number formatting, bold headers, frozen header row.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
import polars as pl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from rollup.schemas.columns import EpType
from rollup.reports import (
    GRAIN_PERIL, GRAIN_ROLLUP_LOB, GRAIN_TOTAL,
    REPORT_EP_TYPE, REPORT_GRAIN, REPORT_GROUP_KEY,
    REPORT_RP, REPORT_VALUE, REPORT_VARIANT,
)


log = logging.getLogger("rollup.report.writer")


_GRAIN_ORDER:  tuple[str, ...] = (GRAIN_TOTAL, GRAIN_ROLLUP_LOB, GRAIN_PERIL)
_EP_ORDER:     tuple[str, ...] = (EpType.AAL, EpType.AEP, EpType.OEP)
_HEADER_FILL = PatternFill(start_color="FF1F2937", end_color="FF1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFFFF")
_NUMBER_FMT  = "#,##0"


def write_report(report: pl.DataFrame, output_dir: Path) -> tuple[Path, Path]:
    """Write `report.csv` and `report.xlsx` under `output_dir`. Returns both paths."""
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path  = output_dir / "report.csv"
    xlsx_path = output_dir / "report.xlsx"

    report.write_csv(csv_path)
    log.info(f"report: wrote {csv_path.name} ({report.height:,} rows)")

    _write_xlsx(report, xlsx_path)
    log.info(f"report: wrote {xlsx_path.name}")
    return csv_path, xlsx_path


def _pivot_for_grain(report: pl.DataFrame, grain: str) -> pl.DataFrame:
    """Wide layout for one grain. Columns: variant, group_key, AAL, AEP_100, OEP_100, ..."""
    grain_df = report.filter(pl.col(REPORT_GRAIN) == grain)
    if grain_df.is_empty():
        return grain_df

    rps = sorted({int(r) for r in grain_df[REPORT_RP].unique().to_list() if r != 0})
    metric_cols: list[str] = [EpType.AAL]
    for rp in rps:
        for ep in (EpType.AEP, EpType.OEP):
            metric_cols.append(_metric_col_name(ep, rp))

    wide = (
        grain_df
        .with_columns(
            pl.when(pl.col(REPORT_EP_TYPE) == EpType.AAL)
              .then(pl.lit(EpType.AAL))
              .otherwise(pl.format("{}_{}", REPORT_EP_TYPE, REPORT_RP))
              .alias("_metric"),
        )
        .pivot(on="_metric", index=[REPORT_VARIANT, REPORT_GROUP_KEY], values=REPORT_VALUE)
    )

    # Ensure every metric column is present, even when one slice didn't produce a row.
    missing = [c for c in metric_cols if c not in wide.columns]
    if missing:
        wide = wide.with_columns([pl.lit(None, dtype=pl.Float64).alias(c) for c in missing])

    return wide.select([REPORT_VARIANT, REPORT_GROUP_KEY, *metric_cols]).sort(REPORT_VARIANT, REPORT_GROUP_KEY)


def _metric_col_name(ep_type: str, rp: int) -> str:
    return f"{ep_type}_{rp}"


def _write_xlsx(report: pl.DataFrame, path: Path) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for grain in _GRAIN_ORDER:
        wide = _pivot_for_grain(report, grain)
        ws = wb.create_sheet(title=grain)

        if wide.is_empty():
            ws["A1"] = f"(no rows for grain={grain})"
            continue

        headers = list(wide.columns)
        ws.append(headers)
        for row in wide.iter_rows():
            ws.append(list(row))

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = Alignment(horizontal="center")

        # Number formatting for value columns (everything after variant+group_key).
        first_value_col = 3  # 1=variant, 2=group_key, 3..N=values
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(first_value_col, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).number_format = _NUMBER_FMT

        # Auto-size columns from content (a cheap heuristic, not perfect).
        for col_idx, header in enumerate(headers, start=1):
            letter = get_column_letter(col_idx)
            width = max(
                len(str(header)),
                *(len(str(ws.cell(row=r, column=col_idx).value)) for r in range(2, min(ws.max_row, 50) + 1)),
            )
            ws.column_dimensions[letter].width = min(max(width + 2, 12), 32)

        ws.freeze_panes = "C2"  # freeze variant + group_key columns and header row

    wb.save(path)
