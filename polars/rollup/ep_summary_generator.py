from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook


logger = logging.getLogger(__name__)

CANONICAL_COLUMNS = [
    "vendor",
    "analysis_id",
    "modelled_lob",
    "modelled_peril",
    "ep_type",
    "return_period",
    "loss",
]


def generate_ep_summaries(data_root: Path | str = "data") -> list[Path]:
    data_root = Path(data_root)
    outputs = [
        _write_ep_summary(
            build_verisk_ep_summary(data_root),
            data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv",
        ),
        _write_ep_summary(
            build_risklink_ep_summary(data_root),
            data_root / "ep_summaries" / "risklink" / "rms_ep_summary.long.csv",
        ),
    ]
    return outputs


def build_verisk_ep_summary(data_root: Path | str = "data") -> pl.DataFrame:
    workbook_path = Path(data_root) / "ep_summaries" / "verisk" / "verisk.xlsx"
    worksheet = load_workbook(workbook_path, read_only=True, data_only=True)["PML by LOB"]
    headers = {value: index for index, value in enumerate(_row_values(worksheet, 7), start=1)}
    metric_columns = [
        (name, column)
        for name, column in headers.items()
        if isinstance(name, str) and (name.startswith("aal_") or name.startswith("aep_") or name.startswith("oep_"))
    ]

    rows: list[dict[str, Any]] = []
    for row_number in range(8, worksheet.max_row + 1):
        analysis = _clean_string(worksheet.cell(row_number, headers["Analysis"]).value)
        modelled_lob = _clean_string(worksheet.cell(row_number, headers["ExposureAttribute"]).value)
        catalog_type = _clean_string(worksheet.cell(row_number, headers["CatalogTypeCode"]).value)
        if not analysis or not modelled_lob or catalog_type != "STC":
            continue

        for metric_name, column in metric_columns:
            loss = worksheet.cell(row_number, column).value
            if loss is None:
                continue
            ep_type, return_period = _parse_verisk_metric(metric_name)
            rows.append(
                {
                    "vendor": "verisk",
                    "analysis_id": analysis,
                    "modelled_lob": modelled_lob,
                    "modelled_peril": analysis,
                    "ep_type": ep_type,
                    "return_period": return_period,
                    "loss": float(loss),
                }
            )

    return _canonical_frame(rows)


def build_risklink_ep_summary(data_root: Path | str = "data") -> pl.DataFrame:
    workbook_path = (
        Path(data_root)
        / "ep_summaries"
        / "risklink"
        / "Hiscox RNL26 RMS by LOB (EDM & RDM).xlsx"
    )
    worksheet = load_workbook(workbook_path, read_only=True, data_only=True)["OEPAEP Curves"]
    header_width = _last_populated_column(worksheet, (5, 6))
    field_columns = {
        _clean_string(worksheet.cell(6, column).value): column
        for column in range(1, header_width + 1)
        if worksheet.cell(6, column).value is not None
    }
    metric_columns = []
    for column in range(1, header_width + 1):
        ep_type = _clean_string(worksheet.cell(5, column).value)
        return_period = worksheet.cell(6, column).value
        if ep_type in {"OEP", "AEP"} and isinstance(return_period, int | float):
            metric_columns.append((ep_type, int(return_period), column))

    rows: list[dict[str, Any]] = []
    for row_number in range(7, worksheet.max_row + 1):
        analysis_id = worksheet.cell(row_number, field_columns["ID"]).value
        modelled_lob = _clean_string(worksheet.cell(row_number, field_columns["LOB"]).value)
        modelled_peril = _clean_string(worksheet.cell(row_number, field_columns["RegionPeril"]).value)
        if analysis_id is None or not modelled_lob or not modelled_peril:
            continue

        aal = worksheet.cell(row_number, field_columns["AAL"]).value
        if aal is not None:
            rows.append(
                {
                    "vendor": "risklink",
                    "analysis_id": str(int(analysis_id)),
                    "modelled_lob": modelled_lob,
                    "modelled_peril": modelled_peril,
                    "ep_type": "AAL",
                    "return_period": 0,
                    "loss": float(aal),
                }
            )

        for ep_type, return_period, column in metric_columns:
            loss = worksheet.cell(row_number, column).value
            if loss is None:
                continue
            rows.append(
                {
                    "vendor": "risklink",
                    "analysis_id": str(int(analysis_id)),
                    "modelled_lob": modelled_lob,
                    "modelled_peril": modelled_peril,
                    "ep_type": ep_type,
                    "return_period": return_period,
                    "loss": float(loss),
                }
            )

    return _canonical_frame(rows)


def _write_ep_summary(frame: pl.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    started = time.perf_counter()
    logger.info("writing output=%s", output_path)
    frame.write_csv(output_path)
    logger.info(
        "wrote output=%s rows=%d elapsed=%.2fs",
        output_path,
        frame.height,
        time.perf_counter() - started,
    )
    return output_path


def _canonical_frame(rows: list[dict[str, Any]]) -> pl.DataFrame:
    return pl.DataFrame(
        rows,
        schema={
            "vendor": pl.String,
            "analysis_id": pl.String,
            "modelled_lob": pl.String,
            "modelled_peril": pl.String,
            "ep_type": pl.String,
            "return_period": pl.Int64,
            "loss": pl.Float64,
        },
    ).select(CANONICAL_COLUMNS)


def _parse_verisk_metric(metric_name: str) -> tuple[str, int]:
    ep_type, value = metric_name.split("_", maxsplit=1)
    if ep_type == "aal":
        return "AAL", 0
    return ep_type.upper(), int(float(value))


def _row_values(worksheet: Any, row_number: int) -> list[Any]:
    return [cell.value for cell in worksheet[row_number]]


def _clean_string(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None


def _last_populated_column(worksheet: Any, row_numbers: tuple[int, ...]) -> int:
    last_column = 1
    for row_number in row_numbers:
        for column, value in enumerate(_row_values(worksheet, row_number), start=1):
            if value is not None:
                last_column = max(last_column, column)
    return last_column
