from __future__ import annotations

import logging
import time
from pathlib import Path
from typing import Any

import polars as pl
from openpyxl import load_workbook


logger = logging.getLogger(__name__)

CANONICAL_COLUMNS = [
    "vendor",
    "analysis_id",
    "modelled_lob",
    "modelled_peril",
    "ep_type",
    "return_period",
    "loss",
]
RISKLINK_HEADER_WIDTH = 30


def generate_ep_summaries(data_root: Path | str = "data") -> list[Path]:
    data_root = Path(data_root)
    outputs = [
        _write_ep_summary(
            build_verisk_ep_summary(data_root),
            data_root / "ep_summaries" / "verisk" / "verisk_ep_summary.long.csv",
        ),
        _write_ep_summary(
            build_risklink_ep_summary(data_root),
            data_root / "ep_summaries" / "risklink" / "rms_ep_summary.long.csv",
        ),
    ]
    return outputs


def build_verisk_ep_summary(data_root: Path | str = "data") -> pl.DataFrame:
    workbook_path = Path(data_root) / "ep_summaries" / "verisk" / "verisk.xlsx"
    worksheet = load_workbook(workbook_path, read_only=True, data_only=True)["PML by LOB"]
    headers = {value: index for index, value in enumerate(_row_values(worksheet, 7), start=1)}
    metric_columns = [
        (name, column)
        for name, column in headers.items()
        if isinstance(name, str) and (name.startswith("aal_") or name.startswith("aep_") or name.startswith("oep_"))
    ]

    rows: list[dict[str, Any]] = []
    for row_number in range(8, worksheet.max_row + 1):
        analysis = _clean_string(worksheet.cell(row_number, headers["Analysis"]).value)
        modelled_lob = _clean_string(worksheet.cell(row_number, headers["ExposureAttribute"]).value)
        catalog_type = _clean_string(worksheet.cell(row_number, headers["CatalogTypeCode"]).value)
        if not analysis or not modelled_lob or catalog_type != "STC":
            continue

        for metric_name, column in metric_columns:
            loss = worksheet.cell(row_number, column).value
            if loss is None:
                continue
            ep_type, return_period = _parse_verisk_metric(metric_name)
            rows.append(
                {
                    "vendor": "verisk",
                    "analysis_id": analysis,
                    "modelled_lob": modelled_lob,
                    "modelled_peril": analysis,
                    "ep_type": ep_type,
                    "return_period": return_period,
                    "loss": float(loss),
                }
            )

    return _canonical_frame(rows)


def build_risklink_ep_summary(data_root: Path | str = "data") -> pl.DataFrame:
    workbook_path = (
        Path(data_root)
        / "ep_summaries"
        / "risklink"
        / "Hiscox RNL26 RMS by LOB (EDM & RDM).xlsx"
    )
    worksheet = load_workbook(workbook_path, read_only=True, data_only=True)["OEPAEP Curves"]
    header_rows = list(
        worksheet.iter_rows(
            min_row=5,
            max_row=6,
            min_col=1,
            max_col=RISKLINK_HEADER_WIDTH,
            values_only=True,
        )
    )
    ep_type_headers = header_rows[0]
    field_headers = header_rows[1]
    field_columns = {
        _clean_string(value): index
        for index, value in enumerate(field_headers)
        if value is not None
    }
    metric_columns = []
    for index, (ep_type_raw, return_period) in enumerate(zip(ep_type_headers, field_headers, strict=True)):
        ep_type = _clean_string(ep_type_raw)
        if ep_type in {"OEP", "AEP"} and isinstance(return_period, int | float):
            metric_columns.append((ep_type, int(return_period), index))

    rows: list[dict[str, Any]] = []
    for row in worksheet.iter_rows(
        min_row=7,
        max_col=RISKLINK_HEADER_WIDTH,
        values_only=True,
    ):
        analysis_id = row[field_columns["ID"]]
        modelled_lob = _clean_string(row[field_columns["LOB"]])
        modelled_peril = _clean_string(row[field_columns["RegionPeril"]])
        if analysis_id is None or not modelled_lob or not modelled_peril:
            continue

        aal = row[field_columns["AAL"]]
        if aal is not None:
            rows.append(
                {
                    "vendor": "risklink",
                    "analysis_id": str(int(analysis_id)),
                    "modelled_lob": modelled_lob,
                    "modelled_peril": modelled_peril,
                    "ep_type": "AAL",
                    "return_period": 0,
                    "loss": float(aal),
                }
            )

        for ep_type, return_period, column in metric_columns:
            loss = row[column]
            if loss is None:
                continue
            rows.append(
                {
                    "vendor": "risklink",
                    "analysis_id": str(int(analysis_id)),
                    "modelled_lob": modelled_lob,
                    "modelled_peril": modelled_peril,
                    "ep_type": ep_type,
                    "return_period": return_period,
                    "loss": float(loss),
                }
            )

    return _canonical_frame(rows)


def _write_ep_summary(frame: pl.DataFrame, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    started = time.perf_counter()
    logger.info("writing output=%s", output_path)
    frame.write_csv(output_path)
    logger.info(
        "wrote output=%s rows=%d elapsed=%.2fs",
        output_path,
        frame.height,
        time.perf_counter() - started,
    )
    return output_path


def _canonical_frame(rows: list[dict[str, Any]]) -> pl.DataFrame:
    return pl.DataFrame(
        rows,
        schema={
            "vendor": pl.String,
            "analysis_id": pl.String,
            "modelled_lob": pl.String,
            "modelled_peril": pl.String,
            "ep_type": pl.String,
            "return_period": pl.Int64,
            "loss": pl.Float64,
        },
    ).select(CANONICAL_COLUMNS)


def _parse_verisk_metric(metric_name: str) -> tuple[str, int]:
    ep_type, value = metric_name.split("_", maxsplit=1)
    if ep_type == "aal":
        return "AAL", 0
    return ep_type.upper(), int(float(value))


def _row_values(worksheet: Any, row_number: int) -> list[Any]:
    return [cell.value for cell in worksheet[row_number]]


def _clean_string(value: Any) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None

